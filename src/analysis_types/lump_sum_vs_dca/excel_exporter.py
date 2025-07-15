"""
Excel 출력 모듈
"""
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1, FORMAT_PERCENTAGE_00
import os
from typing import Dict, Any
from datetime import datetime


class ExcelExporter:
    """Excel 출력 클래스"""
    
    def __init__(self, config):
        self.config = config
        self.workbook = None
        self.filename = None
    
    def export_analysis(self, comparison_result: Dict[str, Any], analyzer) -> str:
        """분석 결과를 Excel로 출력"""
        
        # 파일명 생성
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.filename = f"lump_sum_vs_dca_{self.config.symbol}_{self.config.start_year}{self.config.start_month:02d}_{timestamp}.xlsx"
        filepath = os.path.join(self.config.results_dir, self.filename)
        
        # 워크북 생성
        self.workbook = openpyxl.Workbook()
        
        # 기본 시트 삭제
        self.workbook.remove(self.workbook.active)
        
        # 각 시트 생성
        self._create_purchase_history_sheet(comparison_result)
        self._create_daily_returns_sheet(comparison_result)
        self._create_analysis_summary_sheet(comparison_result, analyzer)
        
        # 파일 저장
        self.workbook.save(filepath)
        
        return filepath
    
    def _create_purchase_history_sheet(self, comparison_result: Dict[str, Any]):
        """매수 내역 시트 생성"""
        ws = self.workbook.create_sheet("매수 내역")
        
        # 헤더 설정
        headers = ['구분', '매수일', '매수가격', '매수금액', '매수수량']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # 데이터 입력
        row = 2
        
        # 일시투자 데이터
        lump_sum_trades = comparison_result['lump_sum']['trades']
        for trade in lump_sum_trades:
            ws.cell(row=row, column=1, value="일시투자")
            ws.cell(row=row, column=2, value=trade['date'])
            ws.cell(row=row, column=3, value=trade['price'])
            ws.cell(row=row, column=4, value=trade['amount'])
            ws.cell(row=row, column=5, value=trade['shares'])
            row += 1
        
        # 적립투자 데이터
        dca_trades = comparison_result['dca']['trades']
        for trade in dca_trades:
            ws.cell(row=row, column=1, value="적립투자")
            ws.cell(row=row, column=2, value=trade['date'])
            ws.cell(row=row, column=3, value=trade['price'])
            ws.cell(row=row, column=4, value=trade['amount'])
            ws.cell(row=row, column=5, value=trade['shares'])
            row += 1
        
        # 셀 서식 적용
        self._apply_cell_formatting(ws)
    
    def _create_daily_returns_sheet(self, comparison_result: Dict[str, Any]):
        """일 수익률 변화 시트 생성"""
        ws = self.workbook.create_sheet("일 수익률 변화")
        
        # 일시투자와 적립투자 데이터 병합
        lump_sum_df = comparison_result['lump_sum']['daily_returns'].copy()
        dca_df = comparison_result['dca']['daily_returns'].copy()
        
        # 컬럼명 변경
        lump_sum_df.columns = [f'일시투자_{col}' if col != 'date' else col for col in lump_sum_df.columns]
        dca_df.columns = [f'적립투자_{col}' if col != 'date' else col for col in dca_df.columns]
        
        # 날짜 기준으로 병합
        merged_df = pd.merge(lump_sum_df, dca_df, on='date', how='outer').sort_values('date')
        
        # 컬럼 순서 정리
        columns_order = [
            'date', 
            '일시투자_price', '일시투자_invested_amount', '일시투자_shares', 
            '일시투자_average_price', '일시투자_current_value', '일시투자_total_return',
            '적립투자_price', '적립투자_invested_amount', '적립투자_shares',
            '적립투자_average_price', '적립투자_current_value', '적립투자_total_return'
        ]
        
        # 존재하는 컬럼만 선택
        available_columns = [col for col in columns_order if col in merged_df.columns]
        merged_df = merged_df[available_columns]
        
        # 컬럼명 한글화
        column_mapping = {
            'date': '날짜',
            '일시투자_price': '일시투자_현재가',
            '일시투자_invested_amount': '일시투자_투자금액',
            '일시투자_shares': '일시투자_보유수량',
            '일시투자_average_price': '일시투자_평균단가',
            '일시투자_current_value': '일시투자_현재가치',
            '일시투자_total_return': '일시투자_수익률',
            '적립투자_price': '적립투자_현재가',
            '적립투자_invested_amount': '적립투자_투자금액',
            '적립투자_shares': '적립투자_보유수량',
            '적립투자_average_price': '적립투자_평균단가',
            '적립투자_current_value': '적립투자_현재가치',
            '적립투자_total_return': '적립투자_수익률'
        }
        
        merged_df.rename(columns=column_mapping, inplace=True)
        
        # 데이터프레임을 시트에 입력
        for r in dataframe_to_rows(merged_df, index=False, header=True):
            ws.append(r)
        
        # 셀 서식 적용
        self._apply_cell_formatting(ws)
    
    def _create_analysis_summary_sheet(self, comparison_result: Dict[str, Any], analyzer):
        """분석 요약 시트 생성"""
        ws = self.workbook.create_sheet("분석 요약")
        
        # 지표 계산
        lump_sum_metrics = analyzer.calculate_metrics(comparison_result['lump_sum'])
        dca_metrics = analyzer.calculate_metrics(comparison_result['dca'])
        
        # 요약 테이블 생성
        summary_data = [
            ['구분', '일시투자', '적립투자', '차이'],
            ['최종 수익률', f"{lump_sum_metrics['final_return']*100:.2f}%", 
             f"{dca_metrics['final_return']*100:.2f}%", 
             f"{(lump_sum_metrics['final_return'] - dca_metrics['final_return'])*100:.2f}%"],
            ['CAGR', f"{lump_sum_metrics['cagr']*100:.2f}%", 
             f"{dca_metrics['cagr']*100:.2f}%", 
             f"{(lump_sum_metrics['cagr'] - dca_metrics['cagr'])*100:.2f}%"],
            ['MDD', f"{lump_sum_metrics['mdd']*100:.2f}%", 
             f"{dca_metrics['mdd']*100:.2f}%", 
             f"{(lump_sum_metrics['mdd'] - dca_metrics['mdd'])*100:.2f}%"],
            ['샤프 지수', f"{lump_sum_metrics['sharpe_ratio']:.3f}", 
             f"{dca_metrics['sharpe_ratio']:.3f}", 
             f"{lump_sum_metrics['sharpe_ratio'] - dca_metrics['sharpe_ratio']:.3f}"],
            ['변동성', f"{lump_sum_metrics['volatility']*100:.2f}%", 
             f"{dca_metrics['volatility']*100:.2f}%", 
             f"{(lump_sum_metrics['volatility'] - dca_metrics['volatility'])*100:.2f}%"],
            ['승률', f"{lump_sum_metrics['win_rate']*100:.2f}%", 
             f"{dca_metrics['win_rate']*100:.2f}%", 
             f"{(lump_sum_metrics['win_rate'] - dca_metrics['win_rate'])*100:.2f}%"],
            ['최종 가치', f"{lump_sum_metrics['final_value']:,.0f}원", 
             f"{dca_metrics['final_value']:,.0f}원", 
             f"{lump_sum_metrics['final_value'] - dca_metrics['final_value']:,.0f}원"]
        ]
        
        # 데이터 입력
        for row_data in summary_data:
            ws.append(row_data)
        
        # 투자 설정 정보 추가
        ws.append([])
        ws.append(['[ 투자 설정 정보 ]'])
        ws.append(['지수', self.config.symbol])
        ws.append(['투자 시작', f"{self.config.start_year}년 {self.config.start_month}월"])
        ws.append(['투자 기간', f"{self.config.investment_period_years}년"])
        ws.append(['적립 분할 월수', f"{self.config.dca_months}개월"])
        ws.append(['총 투자금', f"{self.config.initial_capital:,}원"])
        ws.append(['월 적립금', f"{self.config.get_dca_monthly_amount():,.0f}원"])
        
        # 셀 서식 적용
        self._apply_cell_formatting(ws)
    
    def _apply_cell_formatting(self, ws):
        """셀 서식 적용"""
        # 머리행 고정
        ws.freeze_panes = 'A2'
        
        # 열 너비 조정
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 18)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 데이터 타입별 서식 적용
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                
                # 날짜 형식
                if '날짜' in str(ws.cell(1, cell.column).value) or 'date' in str(ws.cell(1, cell.column).value):
                    cell.alignment = Alignment(horizontal="center")
                
                # 수익률 형식
                elif '수익률' in str(ws.cell(1, cell.column).value) or 'return' in str(ws.cell(1, cell.column).value):
                    if isinstance(cell.value, (int, float)) and cell.row > 1:
                        cell.number_format = FORMAT_PERCENTAGE_00
                        cell.alignment = Alignment(horizontal="right")
                
                # 금액 형식
                elif any(keyword in str(ws.cell(1, cell.column).value) for keyword in ['금액', '가치', '가격', 'amount', 'value', 'price']):
                    if isinstance(cell.value, (int, float)) and cell.row > 1:
                        cell.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
                        cell.alignment = Alignment(horizontal="right")
                
                # 수량 형식
                elif any(keyword in str(ws.cell(1, cell.column).value) for keyword in ['수량', '평균단가', 'shares', 'average_price']):
                    if isinstance(cell.value, (int, float)) and cell.row > 1:
                        cell.number_format = "#,##0.0000"
                        cell.alignment = Alignment(horizontal="right")
                
                # 일반 숫자 형식
                elif isinstance(cell.value, (int, float)) and cell.row > 1:
                    cell.alignment = Alignment(horizontal="right")