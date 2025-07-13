"""
확률 기반 백테스팅 분석 모듈
1972년~2015년 매월 시작점에서 일시투자 vs 적립식투자 성과 비교
"""
from dataclasses import dataclass
from datetime import datetime, timedelta
from typing import Dict, List, Tuple
import pandas as pd
import numpy as np
from dateutil.relativedelta import relativedelta


@dataclass
class ScenarioConfig:
    """시나리오별 투자 설정"""
    total_amount: float = 60000  # 총 투자금액
    investment_period_months: int = 60  # 적립식 투자 기간 (5년)
    analysis_period_years: int = 10  # 성과 분석 기간
    start_year: int = 1972
    start_month: int = 1
    end_year: int = 2015
    end_month: int = 1


@dataclass
class ScenarioResult:
    """개별 시나리오 결과"""
    start_date: str
    end_date: str
    lump_sum_return: float  # 일시투자 수익률 (%)
    dca_return: float  # 적립식투자 수익률 (%)
    lump_sum_final_value: float  # 일시투자 최종 가치
    dca_final_value: float  # 적립식투자 최종 가치
    winner: str  # "일시투자" or "적립식투자"
    return_difference: float  # 수익률 차이 (%p)
    # 새로 추가된 지표들
    lump_sum_cagr: float  # 일시투자 연평균수익률 (%)
    dca_cagr: float  # 적립식투자 연평균수익률 (%)
    lump_sum_mdd: float  # 일시투자 최대낙폭 (%)
    dca_mdd: float  # 적립식투자 최대낙폭 (%)
    lump_sum_sharpe: float  # 일시투자 샤프지수
    dca_sharpe: float  # 적립식투자 샤프지수
    dca_avg_price: float  # 적립식투자 평균단가
    dca_total_shares: float  # 적립식투자 총 구매 수량
    start_price: float  # 투자 시작일 지수 가격
    end_price: float  # 투자 종료일 지수 가격
    
    def to_dict(self) -> Dict:
        """딕셔너리로 변환"""
        return {
            "시작일": self.start_date,
            "측정일": self.end_date,
            "일시투자_수익률": round(self.lump_sum_return, 2),
            "적립식투자_수익률": round(self.dca_return, 2),
            "일시투자_최종가치": round(self.lump_sum_final_value, 0),
            "적립식투자_최종가치": round(self.dca_final_value, 0),
            "승자": self.winner,
            "수익률차이": round(self.return_difference, 2),
            "일시투자_CAGR": round(self.lump_sum_cagr, 2),
            "적립식투자_CAGR": round(self.dca_cagr, 2),
            "일시투자_MDD": round(self.lump_sum_mdd, 2),
            "적립식투자_MDD": round(self.dca_mdd, 2),
            "일시투자_샤프지수": round(self.lump_sum_sharpe, 3),
            "적립식투자_샤프지수": round(self.dca_sharpe, 3),
            "적립식_평단가": round(self.dca_avg_price, 2),
            "적립식_총구매수량": round(self.dca_total_shares, 2),
            "시작일_지수가격": round(self.start_price, 2),
            "측정일_지수가격": round(self.end_price, 2)
        }


class ProbabilisticBacktester:
    """확률 기반 백테스팅 엔진"""
    
    def __init__(self, config: ScenarioConfig):
        self.config = config
        self.scenarios: List[ScenarioResult] = []
        
    def generate_scenarios(self) -> List[Tuple[datetime, datetime]]:
        """시나리오 시작/종료 날짜 생성"""
        import pytz
        scenarios = []
        
        current_date = datetime(self.config.start_year, self.config.start_month, 1)
        end_limit = datetime(self.config.end_year, self.config.end_month, 1)
        
        while current_date < end_limit:
            start_date = current_date
            end_date = start_date + relativedelta(years=self.config.analysis_period_years)
            
            # UTC timezone 추가
            start_date_utc = start_date.replace(tzinfo=pytz.UTC)
            end_date_utc = end_date.replace(tzinfo=pytz.UTC)
            
            scenarios.append((start_date_utc, end_date_utc))
            
            # 다음 달로 이동
            current_date = current_date + relativedelta(months=1)
            
        return scenarios
    
    def run_single_scenario(self, price_data: pd.DataFrame, 
                           start_date: datetime, 
                           end_date: datetime) -> ScenarioResult:
        """단일 시나리오 백테스팅 실행"""
        
        # 날짜 범위 필터링
        start_date_str = start_date.strftime('%Y-%m-%d')
        end_date_str = end_date.strftime('%Y-%m-%d')
        
        # 시나리오 날짜는 이미 UTC timezone이 있으므로 직접 비교
        mask = (price_data.index >= start_date) & (price_data.index <= end_date)
        
        scenario_data = price_data[mask].copy()
        
        if len(scenario_data) < 120:  # 최소 10년 데이터 필요
            raise ValueError(f"데이터 부족: {start_date_str} ~ {end_date_str}")
        
        # 일시투자 백테스팅
        lump_sum_result = self._run_lump_sum(scenario_data)
        
        # 적립식투자 백테스팅
        dca_result = self._run_dca(scenario_data)
        
        # 결과 비교
        winner = "일시투자" if lump_sum_result['return'] > dca_result['return'] else "적립식투자"
        return_diff = lump_sum_result['return'] - dca_result['return']
        
        # 시작가격과 종료가격 추출
        start_price = scenario_data['Open'].iloc[0]
        end_price = scenario_data['Close'].iloc[-1]
        
        return ScenarioResult(
            start_date=start_date_str,
            end_date=end_date_str,
            lump_sum_return=lump_sum_result['return'],
            dca_return=dca_result['return'],
            lump_sum_final_value=lump_sum_result['final_value'],
            dca_final_value=dca_result['final_value'],
            winner=winner,
            return_difference=return_diff,
            lump_sum_cagr=lump_sum_result['cagr'],
            dca_cagr=dca_result['cagr'],
            lump_sum_mdd=lump_sum_result['mdd'],
            dca_mdd=dca_result['mdd'],
            lump_sum_sharpe=lump_sum_result['sharpe'],
            dca_sharpe=dca_result['sharpe'],
            dca_avg_price=dca_result['avg_price'],
            dca_total_shares=dca_result['shares'],
            start_price=start_price,
            end_price=end_price
        )
    
    def _run_lump_sum(self, data: pd.DataFrame) -> Dict:
        """일시투자 백테스팅"""
        initial_price = data['Open'].iloc[0]
        final_price = data['Close'].iloc[-1]
        
        shares = self.config.total_amount / initial_price
        final_value = shares * final_price
        total_return = (final_value / self.config.total_amount - 1) * 100
        
        # CAGR 계산 (연평균 수익률)
        years = self.config.analysis_period_years
        cagr = ((final_value / self.config.total_amount) ** (1/years) - 1) * 100 if years > 0 else 0
        
        # MDD 및 샤프지수 계산을 위한 가격 시계열 생성
        portfolio_values = data['Close'] * shares
        daily_returns = portfolio_values.pct_change().dropna()
        
        # MDD 계산
        cumulative_returns = portfolio_values / portfolio_values.iloc[0]
        rolling_max = cumulative_returns.expanding().max()
        drawdown = (cumulative_returns - rolling_max) / rolling_max * 100
        mdd = drawdown.min()
        
        # 샤프지수 계산
        if len(daily_returns) > 0 and daily_returns.std() > 0:
            excess_return = cagr - 2.0  # 무위험 수익률 2% 가정
            volatility = daily_returns.std() * np.sqrt(252) * 100  # 연환산 변동성
            sharpe_ratio = excess_return / volatility if volatility > 0 else 0
        else:
            sharpe_ratio = 0
        
        return {
            'final_value': final_value,
            'return': total_return,
            'shares': shares,
            'cagr': cagr,
            'mdd': mdd,
            'sharpe': sharpe_ratio
        }
    
    def _run_dca(self, data: pd.DataFrame) -> Dict:
        """적립식투자 백테스팅 (5년간 매월 투자)"""
        monthly_amount = self.config.total_amount / self.config.investment_period_months
        total_shares = 0
        total_invested = 0
        investment_records = []  # 투자 기록 저장
        
        # 매월 투자 (60개월)
        for month in range(min(self.config.investment_period_months, len(data))):
            if month < len(data):
                monthly_price = data['Open'].iloc[month]
                shares_bought = monthly_amount / monthly_price
                total_shares += shares_bought
                total_invested += monthly_amount
                investment_records.append({
                    'month': month,
                    'price': monthly_price,
                    'amount': monthly_amount,
                    'shares': shares_bought
                })
        
        # 평단가 계산
        avg_price = total_invested / total_shares if total_shares > 0 else 0
        
        # 최종 가치 계산 (10년 후)
        final_price = data['Close'].iloc[-1]
        final_value = total_shares * final_price
        total_return = (final_value / total_invested - 1) * 100
        
        # CAGR 계산
        years = self.config.analysis_period_years
        cagr = ((final_value / total_invested) ** (1/years) - 1) * 100 if years > 0 and total_invested > 0 else 0
        
        # MDD 및 샤프지수 계산을 위한 포트폴리오 가치 시계열 생성
        portfolio_values = pd.Series(index=data.index, dtype=float)
        current_shares = 0
        current_invested = 0
        investment_month = 0
        
        for i, date in enumerate(data.index):
            # 투자 기간 중이면 매월 투자 진행
            if investment_month < self.config.investment_period_months and i % 21 == 0:  # 대략 월별 (21 영업일)
                if investment_month < len(investment_records):
                    current_shares += investment_records[investment_month]['shares']
                    current_invested += investment_records[investment_month]['amount']
                    investment_month += 1
            
            # 포트폴리오 가치 계산
            portfolio_values.iloc[i] = current_shares * data['Close'].iloc[i]
        
        # 실제 투자한 부분만 사용하여 수익률 계산
        portfolio_values = portfolio_values.dropna()
        if len(portfolio_values) > 1:
            # 포트폴리오 가치 기준 수익률 계산 (원금 대비)
            invested_series = pd.Series(index=portfolio_values.index)
            month_tracker = 0
            invested_amount = 0
            
            for i, date in enumerate(portfolio_values.index):
                if month_tracker < self.config.investment_period_months and i % 21 == 0:
                    if month_tracker < len(investment_records):
                        invested_amount += investment_records[month_tracker]['amount']
                        month_tracker += 1
                invested_series.iloc[i] = invested_amount if invested_amount > 0 else investment_records[0]['amount']
            
            # MDD 계산 (투자원금 대비)
            returns_ratio = portfolio_values / invested_series
            rolling_max = returns_ratio.expanding().max()
            drawdown = (returns_ratio - rolling_max) / rolling_max * 100
            mdd = drawdown.min()
            
            # 일별 수익률 계산
            daily_returns = portfolio_values.pct_change().dropna()
            
            # 샤프지수 계산
            if len(daily_returns) > 0 and daily_returns.std() > 0:
                excess_return = cagr - 2.0  # 무위험 수익률 2% 가정
                volatility = daily_returns.std() * np.sqrt(252) * 100  # 연환산 변동성
                sharpe_ratio = excess_return / volatility if volatility > 0 else 0
            else:
                sharpe_ratio = 0
        else:
            mdd = 0
            sharpe_ratio = 0
        
        return {
            'final_value': final_value,
            'return': total_return,
            'total_invested': total_invested,
            'shares': total_shares,
            'avg_price': avg_price,
            'cagr': cagr,
            'mdd': mdd,
            'sharpe': sharpe_ratio
        }
    
    def run_all_scenarios(self, price_data: pd.DataFrame) -> None:
        """모든 시나리오 실행"""
        scenarios = self.generate_scenarios()
        self.scenarios = []
        
        print(f"📊 총 {len(scenarios)}개 시나리오 분석 시작...")
        
        successful_scenarios = 0
        failed_scenarios = 0
        
        for i, (start_date, end_date) in enumerate(scenarios):
            try:
                result = self.run_single_scenario(price_data, start_date, end_date)
                self.scenarios.append(result)
                successful_scenarios += 1
                
                # 진행상황 출력 (매 50개마다)
                if (i + 1) % 50 == 0:
                    print(f"  📈 진행상황: {i + 1}/{len(scenarios)} ({(i+1)/len(scenarios)*100:.1f}%)")
                    
            except Exception as e:
                failed_scenarios += 1
                if failed_scenarios <= 5:  # 처음 5개 오류만 출력
                    print(f"  ⚠️ 시나리오 실패: {start_date.strftime('%Y-%m')} - {str(e)}")
        
        print(f"✅ 분석 완료: 성공 {successful_scenarios}개, 실패 {failed_scenarios}개")
    
    def get_summary_statistics(self) -> Dict:
        """요약 통계 생성"""
        if not self.scenarios:
            return {}
        
        # 기본 통계
        total_scenarios = len(self.scenarios)
        lump_sum_wins = sum(1 for s in self.scenarios if s.winner == "일시투자")
        dca_wins = total_scenarios - lump_sum_wins
        
        # 수익률 데이터
        lump_sum_returns = [s.lump_sum_return for s in self.scenarios]
        dca_returns = [s.dca_return for s in self.scenarios]
        return_differences = [s.return_difference for s in self.scenarios]
        
        # 통계 계산
        stats = {
            "기본_통계": {
                "총_시나리오수": total_scenarios,
                "일시투자_승리": lump_sum_wins,
                "적립식투자_승리": dca_wins,
                "일시투자_승률": round(lump_sum_wins / total_scenarios * 100, 1),
                "적립식투자_승률": round(dca_wins / total_scenarios * 100, 1)
            },
            "수익률_통계": {
                "일시투자_평균수익률": round(np.mean(lump_sum_returns), 2),
                "적립식투자_평균수익률": round(np.mean(dca_returns), 2),
                "평균_수익률차이": round(np.mean(return_differences), 2),
                "일시투자_표준편차": round(np.std(lump_sum_returns), 2),
                "적립식투자_표준편차": round(np.std(dca_returns), 2)
            },
            "극값_분석": {
                "일시투자_최고수익률": round(max(lump_sum_returns), 2),
                "일시투자_최저수익률": round(min(lump_sum_returns), 2),
                "적립식투자_최고수익률": round(max(dca_returns), 2),
                "적립식투자_최저수익률": round(min(dca_returns), 2),
                "최대_수익률차이": round(max(return_differences), 2),
                "최소_수익률차이": round(min(return_differences), 2)
            }
        }
        
        # 최고/최악 시나리오 찾기
        best_lump_sum = max(self.scenarios, key=lambda x: x.lump_sum_return)
        worst_lump_sum = min(self.scenarios, key=lambda x: x.lump_sum_return)
        best_dca = max(self.scenarios, key=lambda x: x.dca_return)
        worst_dca = min(self.scenarios, key=lambda x: x.dca_return)
        
        stats["극값_시나리오"] = {
            "일시투자_최고": {
                "시작일": best_lump_sum.start_date,
                "수익률": best_lump_sum.lump_sum_return
            },
            "일시투자_최악": {
                "시작일": worst_lump_sum.start_date,
                "수익률": worst_lump_sum.lump_sum_return
            },
            "적립식투자_최고": {
                "시작일": best_dca.start_date,
                "수익률": best_dca.dca_return
            },
            "적립식투자_최악": {
                "시작일": worst_dca.start_date,
                "수익률": worst_dca.dca_return
            }
        }
        
        return stats
    
    
    def export_to_excel(self, output_path: str = "results/") -> str:
        """결과를 Excel 파일로 내보내기 (다중 시트)"""
        import os
        from datetime import datetime
        
        if not self.scenarios:
            raise ValueError("분석된 시나리오가 없습니다.")
        
        # 결과 디렉토리 생성
        os.makedirs(output_path, exist_ok=True)
        
        # 파일명 생성
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"확률분석_나스닥_{timestamp}.xlsx"
        filepath = os.path.join(output_path, filename)
        
        # 데이터프레임 생성
        df = pd.DataFrame([scenario.to_dict() for scenario in self.scenarios])
        
        # 통계 데이터 생성
        stats = self.get_summary_statistics()
        
        # Excel Writer 생성
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # 1. 전체 시나리오 시트 (수익률을 소수로 변환)
            df_excel = self._prepare_data_for_excel(df)
            df_excel.to_excel(writer, sheet_name='전체시나리오', index=False)
            
            # 2. 전체 통계 시트
            self._create_statistics_sheet(writer, stats)
            
            # 3. 분석 리포트 시트
            self._create_analysis_sheet(writer, df, stats)
            
            # 4. 연도별 통계 시트
            self._create_chart_data_sheet(writer, df)
        
        # 엑셀 파일 서식 적용
        self._format_excel_file(filepath)
        
        return filepath
    
    def _prepare_data_for_excel(self, df: pd.DataFrame) -> pd.DataFrame:
        """Excel용으로 데이터 변환 (수익률을 소수로 변환)"""
        df_excel = df.copy()
        
        # 수익률 컬럼들을 소수로 변환 (Excel에서 %포맷 적용하기 위해)
        percentage_columns = [
            '일시투자_수익률',
            '적립식투자_수익률', 
            '수익률차이',
            '일시투자_CAGR',
            '적립식투자_CAGR',
            '일시투자_MDD',
            '적립식투자_MDD'
        ]
        
        for col in percentage_columns:
            if col in df_excel.columns:
                df_excel[col] = df_excel[col] / 100  # %를 소수로 변환
        
        return df_excel
    
    def _prepare_chart_data_for_excel(self, df: pd.DataFrame) -> pd.DataFrame:
        """차트데이터용 Excel 변환 (수익률을 소수로 변환)"""
        df_excel = df.copy()
        
        # 수익률 관련 컬럼들을 소수로 변환
        percentage_columns = []
        for col in df_excel.columns:
            if any(keyword in col for keyword in ['수익률', 'CAGR', 'MDD', '승률']):
                percentage_columns.append(col)
        
        for col in percentage_columns:
            if col in df_excel.columns:
                df_excel[col] = df_excel[col] / 100  # %를 소수로 변환
        
        return df_excel
    
    def _create_statistics_sheet(self, writer, stats: Dict) -> None:
        """통계 요약 시트 생성"""
        stats_data = []
        
        # 기본 통계
        basic_stats = stats.get('기본_통계', {})
        stats_data.extend([
            ['구분', '항목', '값'],
            ['', '', ''],
            ['기본 통계', '총 시나리오 수', basic_stats.get('총_시나리오수', 0)],
            ['', '일시투자 승리', f"{basic_stats.get('일시투자_승리', 0)}회"],
            ['', '적립식투자 승리', f"{basic_stats.get('적립식투자_승리', 0)}회"],
            ['', '일시투자 승률', f"{basic_stats.get('일시투자_승률', 0)}%"],
            ['', '적립식투자 승률', f"{basic_stats.get('적립식투자_승률', 0)}%"],
            ['', '', '']
        ])
        
        # 수익률 통계
        return_stats = stats.get('수익률_통계', {})
        stats_data.extend([
            ['수익률 통계', '일시투자 평균수익률', f"{return_stats.get('일시투자_평균수익률', 0)}%"],
            ['', '적립식투자 평균수익률', f"{return_stats.get('적립식투자_평균수익률', 0)}%"],
            ['', '평균 수익률 차이', f"{return_stats.get('평균_수익률차이', 0)}%p"],
            ['', '일시투자 표준편차', f"{return_stats.get('일시투자_표준편차', 0)}%"],
            ['', '적립식투자 표준편차', f"{return_stats.get('적립식투자_표준편차', 0)}%"],
            ['', '', '']
        ])
        
        # 극값 분석
        extreme_stats = stats.get('극값_분석', {})
        stats_data.extend([
            ['극값 분석', '일시투자 최고수익률', f"{extreme_stats.get('일시투자_최고수익률', 0)}%"],
            ['', '일시투자 최저수익률', f"{extreme_stats.get('일시투자_최저수익률', 0)}%"],
            ['', '적립식투자 최고수익률', f"{extreme_stats.get('적립식투자_최고수익률', 0)}%"],
            ['', '적립식투자 최저수익률', f"{extreme_stats.get('적립식투자_최저수익률', 0)}%"],
            ['', '최대 수익률 차이', f"{extreme_stats.get('최대_수익률차이', 0)}%p"],
            ['', '최소 수익률 차이', f"{extreme_stats.get('최소_수익률차이', 0)}%p"]
        ])
        
        # DataFrame으로 변환 후 저장
        stats_df = pd.DataFrame(stats_data)
        stats_df.to_excel(writer, sheet_name='전체통계', index=False, header=False)
    
    def _create_analysis_sheet(self, writer, df: pd.DataFrame, stats: Dict) -> None:
        """분석 요약 시트 생성"""
        analysis_data = []
        
        # 제목
        analysis_data.extend([
            ['[분석] 확률 기반 투자 전략 분석 리포트'],
            ['-' * 50],
            [''],
            ['[발견] 주요 발견사항'],
            ['']
        ])
        
        # 승률 분석
        basic_stats = stats.get('기본_통계', {})
        ls_win_rate = basic_stats.get('일시투자_승률', 0)
        dca_win_rate = basic_stats.get('적립식투자_승률', 0)
        
        analysis_data.extend([
            [f"- 일시투자가 {ls_win_rate}%의 확률로 적립식투자보다 우수한 성과"],
            [f"- 적립식투자가 {dca_win_rate}%의 확률로 일시투자보다 우수한 성과"],
            ['']
        ])
        
        # 수익률 분석
        return_stats = stats.get('수익률_통계', {})
        ls_avg = return_stats.get('일시투자_평균수익률', 0)
        dca_avg = return_stats.get('적립식투자_평균수익률', 0)
        diff = return_stats.get('평균_수익률차이', 0)
        
        analysis_data.extend([
            ['[수익률] 평균 수익률 비교'],
            [f"- 일시투자 평균: {ls_avg}%"],
            [f"- 적립식투자 평균: {dca_avg}%"],
            [f"- 차이: {diff}%p"],
            ['']
        ])
        
        # CAGR 분석
        ls_cagr_avg = df['일시투자_CAGR'].mean()
        dca_cagr_avg = df['적립식투자_CAGR'].mean()
        
        analysis_data.extend([
            ['[CAGR] 연평균 수익률 비교'],
            [f"- 일시투자 평균 CAGR: {ls_cagr_avg:.2f}%"],
            [f"- 적립식투자 평균 CAGR: {dca_cagr_avg:.2f}%"],
            ['']
        ])
        
        # MDD 분석
        ls_mdd_avg = df['일시투자_MDD'].mean()
        dca_mdd_avg = df['적립식투자_MDD'].mean()
        
        analysis_data.extend([
            ['[위험] 최대낙폭(MDD) 비교'],
            [f"- 일시투자 평균 MDD: {ls_mdd_avg:.2f}%"],
            [f"- 적립식투자 평균 MDD: {dca_mdd_avg:.2f}%"],
            ['']
        ])
        
        # 샤프지수 분석
        ls_sharpe_avg = df['일시투자_샤프지수'].mean()
        dca_sharpe_avg = df['적립식투자_샤프지수'].mean()
        
        analysis_data.extend([
            ['[샤프] 샤프지수 비교'],
            [f"- 일시투자 평균 샤프지수: {ls_sharpe_avg:.3f}"],
            [f"- 적립식투자 평균 샤프지수: {dca_sharpe_avg:.3f}"],
            ['']
        ])
        
        # 결론
        better_strategy = "일시투자" if ls_win_rate > 50 else "적립식투자"
        analysis_data.extend([
            ['[결론] 종합 결론'],
            [f"- 분석 기간 동안 {better_strategy}가 더 우수한 성과를 보임"],
            [f"- 하지만 시장 상황에 따라 결과가 달라질 수 있음"],
            [f"- 투자자의 위험 성향과 투자 목표를 고려한 선택 필요"]
        ])
        
        # DataFrame으로 변환 후 저장
        analysis_df = pd.DataFrame(analysis_data)
        analysis_df.to_excel(writer, sheet_name='분석리포트', index=False, header=False)
    
    def _create_chart_data_sheet(self, writer, df: pd.DataFrame) -> None:
        """차트용 데이터 시트 생성"""
        # 연도별 집계
        df_copy = df.copy()
        df_copy['연도'] = pd.to_datetime(df_copy['시작일']).dt.year
        
        # 기본 통계 계산
        yearly_basic = df_copy.groupby('연도').agg({
            '시작일': 'count',  # 시나리오 수
            '일시투자_수익률': ['mean', 'std', 'min', 'max'],
            '적립식투자_수익률': ['mean', 'std', 'min', 'max'],
            '일시투자_CAGR': 'mean',
            '적립식투자_CAGR': 'mean',
            '일시투자_MDD': 'mean',
            '적립식투자_MDD': 'mean',
            '일시투자_샤프지수': 'mean',
            '적립식투자_샤프지수': 'mean'
        })
        
        # 컬럼명 정리
        yearly_basic.columns = [
            '시나리오수',
            '일시투자_평균수익률', '일시투자_수익률_표준편차', '일시투자_최저수익률', '일시투자_최고수익률',
            '적립투자_평균수익률', '적립투자_수익률_표준편차', '적립투자_최저수익률', '적립투자_최고수익률',
            '일시투자_CAGR', '적립투자_CAGR',
            '일시투자_MDD', '적립투자_MDD',
            '일시투자_샤프지수', '적립투자_샤프지수'
        ]
        
        # 승률 계산
        yearly_wins = df_copy.groupby('연도')['승자'].value_counts().unstack(fill_value=0)
        if '일시투자' in yearly_wins.columns and '적립식투자' in yearly_wins.columns:
            total_scenarios = yearly_wins['일시투자'] + yearly_wins['적립식투자']
            yearly_wins['일시투자_승률'] = (yearly_wins['일시투자'] / total_scenarios * 100).round(1)
            yearly_wins['적립투자_승률'] = (yearly_wins['적립식투자'] / total_scenarios * 100).round(1)
        else:
            yearly_wins['일시투자_승률'] = 0
            yearly_wins['적립투자_승률'] = 0
        
        # 데이터 결합
        chart_data = yearly_basic.join(yearly_wins[['일시투자_승률', '적립투자_승률']], how='left')
        
        # 컬럼 순서 재정렬 (핵심 지표 우선)
        column_order = [
            '시나리오수',
            '일시투자_승률', '적립투자_승률',
            '일시투자_평균수익률', '적립투자_평균수익률',
            '일시투자_CAGR', '적립투자_CAGR',
            '일시투자_MDD', '적립투자_MDD',
            '일시투자_샤프지수', '적립투자_샤프지수',
            '일시투자_수익률_표준편차', '적립투자_수익률_표준편차',
            '일시투자_최고수익률', '일시투자_최저수익률',
            '적립투자_최고수익률', '적립투자_최저수익률'
        ]
        
        # 존재하는 컬럼만 선택
        available_columns = [col for col in column_order if col in chart_data.columns]
        chart_data = chart_data[available_columns]
        
        # 소수점 정리
        chart_data = chart_data.round(2)
        
        # Excel용 포맷 변환 (수익률을 소수로)
        chart_data_excel = self._prepare_chart_data_for_excel(chart_data)
        
        # 저장
        chart_data_excel.to_excel(writer, sheet_name='연도별통계')
    
    def _format_excel_file(self, filepath: str) -> None:
        """Excel 파일 서식 적용"""
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle
            from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00, FORMAT_NUMBER_COMMA_SEPARATED1
            
            wb = load_workbook(filepath)
            
            # 전체시나리오 시트 서식
            if '전체시나리오' in wb.sheetnames:
                self._format_detail_sheet(wb['전체시나리오'])
            
            # 전체통계 시트 서식
            if '전체통계' in wb.sheetnames:
                self._format_statistics_sheet(wb['전체통계'])
            
            # 분석리포트 시트 서식
            if '분석리포트' in wb.sheetnames:
                self._format_analysis_sheet(wb['분석리포트'])
            
            # 연도별통계 시트 서식
            if '연도별통계' in wb.sheetnames:
                self._format_chart_data_sheet(wb['연도별통계'])
            
            wb.save(filepath)
            
        except Exception as e:
            print(f"⚠️ Excel 서식 적용 중 오류: {e}")
            # 서식 적용 실패해도 파일은 저장됨
    
    def _format_detail_sheet(self, ws) -> None:
        """상세데이터 시트 서식 적용"""
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00, FORMAT_NUMBER_COMMA_SEPARATED1
        
        # 헤더 스타일
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # 헤더 행 서식 적용
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 머릿말 행 고정
        ws.freeze_panes = 'A2'
        
        # 컬럼별 데이터 타입에 맞는 서식 적용
        column_formats = {
            'A': None,  # 시작일 (날짜)
            'B': None,  # 측정일 (날짜)
            'C': FORMAT_PERCENTAGE_00,  # 일시투자_수익률
            'D': FORMAT_PERCENTAGE_00,  # 적립식투자_수익률
            'E': FORMAT_NUMBER_COMMA_SEPARATED1,  # 일시투자_최종가치
            'F': FORMAT_NUMBER_COMMA_SEPARATED1,  # 적립식투자_최종가치
            'G': None,  # 승자
            'H': FORMAT_PERCENTAGE_00,  # 수익률차이
            'I': FORMAT_PERCENTAGE_00,  # 일시투자_CAGR
            'J': FORMAT_PERCENTAGE_00,  # 적립식투자_CAGR
            'K': FORMAT_PERCENTAGE_00,  # 일시투자_MDD
            'L': FORMAT_PERCENTAGE_00,  # 적립식투자_MDD
            'M': '#,##0.000',  # 일시투자_샤프지수
            'N': '#,##0.000',  # 적립식투자_샤프지수
            'O': FORMAT_NUMBER_COMMA_SEPARATED1,  # 적립식_평단가
            'P': '#,##0.00',  # 적립식_총구매수량
            'Q': FORMAT_NUMBER_COMMA_SEPARATED1,  # 시작일_지수가격
            'R': FORMAT_NUMBER_COMMA_SEPARATED1,  # 측정일_지수가격
        }
        
        # 데이터 행에 서식 적용
        for row_num in range(2, ws.max_row + 1):
            for col_letter, number_format in column_formats.items():
                if number_format:
                    cell = ws[f'{col_letter}{row_num}']
                    cell.number_format = number_format
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell = ws[f'{col_letter}{row_num}']
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 컬럼 너비 자동 조정 (개선된 로직)
        column_widths = {
            'A': 12,  # 시작일
            'B': 12,  # 측정일
            'C': 14,  # 일시투자_수익률
            'D': 14,  # 적립식투자_수익률
            'E': 16,  # 일시투자_최종가치
            'F': 16,  # 적립식투자_최종가치
            'G': 10,  # 승자
            'H': 12,  # 수익률차이
            'I': 14,  # 일시투자_CAGR
            'J': 14,  # 적립식투자_CAGR
            'K': 14,  # 일시투자_MDD
            'L': 14,  # 적립식투자_MDD
            'M': 16,  # 일시투자_샤프지수
            'N': 16,  # 적립식투자_샤프지수
            'O': 14,  # 적립식_평단가
            'P': 16,  # 적립식_총구매수량
            'Q': 16,  # 시작일_지수가격
            'R': 16,  # 측정일_지수가격
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
    
    def _format_statistics_sheet(self, ws) -> None:
        """통계요약 시트 서식 적용"""
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # 제목 헤더 스타일
        title_font = Font(bold=True, size=12, color='FFFFFF')
        title_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        # 구분 헤더 스타일
        category_font = Font(bold=True, size=11)
        category_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        
        # 첫 번째 행 (헤더) 서식
        for cell in ws[1]:
            cell.font = title_font
            cell.fill = title_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 구분별 서식 적용
        for row in ws.iter_rows():
            if row[0].value in ['기본 통계', '수익률 통계', '극값 분석']:
                for cell in row:
                    cell.font = category_font
                    cell.fill = category_fill
                    cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                for cell in row:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # 컬럼 너비 조정
        ws.column_dimensions['A'].width = 15  # 구분
        ws.column_dimensions['B'].width = 25  # 항목
        ws.column_dimensions['C'].width = 15  # 값
        
        # 머릿말 행 고정
        ws.freeze_panes = 'A2'
    
    def _format_analysis_sheet(self, ws) -> None:
        """분석요약 시트 서식 적용"""
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # 제목 스타일
        title_font = Font(bold=True, size=14, color='366092')
        section_font = Font(bold=True, size=12, color='366092')
        
        # 첫 번째 행 (제목) 서식
        if ws['A1'].value:
            ws['A1'].font = title_font
            ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        
        # 섹션 제목 찾아서 서식 적용
        for row in ws.iter_rows():
            cell_value = str(row[0].value or '')
            if any(keyword in cell_value for keyword in ['[발견]', '[수익률]', '[CAGR]', '[위험]', '[샤프]', '[결론]']):
                row[0].font = section_font
                row[0].alignment = Alignment(horizontal='left', vertical='center')
            else:
                row[0].alignment = Alignment(horizontal='left', vertical='center')
        
        # 컬럼 너비 조정
        ws.column_dimensions['A'].width = 80
    
    def _format_chart_data_sheet(self, ws) -> None:
        """차트데이터 시트 서식 적용"""
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00, FORMAT_NUMBER_COMMA_SEPARATED1
        
        # 헤더 스타일
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        # 헤더 행 서식
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 컬럼별 포맷 정의 (새로운 구조에 맞게)
        column_formats = {}
        for col_num in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_num)
            header_text = str(cell.value or '')
            
            # 연도 컬럼 (첫 번째 컬럼)은 포맷하지 않음
            if col_num == 1:  # 연도 컬럼
                column_formats[col_num] = None
            elif '시나리오수' in header_text:
                column_formats[col_num] = '#,##0'
            elif '승률' in header_text:
                column_formats[col_num] = FORMAT_PERCENTAGE_00
            elif any(keyword in header_text for keyword in ['수익률', 'CAGR', 'MDD']):
                column_formats[col_num] = FORMAT_PERCENTAGE_00
            elif '샤프지수' in header_text:
                column_formats[col_num] = '#,##0.000'
            elif '표준편차' in header_text:
                column_formats[col_num] = FORMAT_PERCENTAGE_00
            else:
                column_formats[col_num] = '#,##0.00'
        
        # 데이터 행에 서식 적용
        for row_num in range(2, ws.max_row + 1):
            for col_num in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_num, column=col_num)
                
                # 포맷 적용 (연도 컬럼은 제외)
                if col_num in column_formats and column_formats[col_num] is not None:
                    cell.number_format = column_formats[col_num]
                
                # 정렬
                if col_num == 1:  # 연도 컬럼
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # 컬럼 너비 설정 (최적화)
        column_widths = {
            1: 8,   # 연도
            2: 12,  # 시나리오수
            3: 12,  # 일시투자_승률
            4: 12,  # 적립투자_승률
            5: 16,  # 일시투자_평균수익률
            6: 16,  # 적립투자_평균수익률
            7: 14,  # 일시투자_CAGR
            8: 14,  # 적립투자_CAGR
            9: 14,  # 일시투자_MDD
            10: 14, # 적립투자_MDD
            11: 16, # 일시투자_샤프지수
            12: 16, # 적립투자_샤프지수
        }
        
        # 나머지 컬럼들은 기본 너비
        for col_num in range(1, ws.max_column + 1):
            width = column_widths.get(col_num, 15)
            col_letter = ws.cell(row=1, column=col_num).column_letter
            ws.column_dimensions[col_letter].width = width
        
        # 머릿말 행 고정
        ws.freeze_panes = 'A2'
    
    def get_scenarios_data(self) -> List[Dict]:
        """시나리오 데이터 리스트 반환"""
        return [scenario.to_dict() for scenario in self.scenarios]