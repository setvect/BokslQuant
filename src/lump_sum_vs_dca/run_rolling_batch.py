"""
롤링 백테스트 배치 실행 (단계별 실행)
전체 522개를 50개씩 나누어 실행
"""

import os
import sys
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from datetime import datetime
from dateutil.relativedelta import relativedelta
from typing import Dict, Any
from pathlib import Path
from rolling_chart_generator import RollingChartGenerator

# 경로 설정
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_src_dir = os.path.dirname(current_dir)
strategies_dir = os.path.join(current_dir, 'strategies')
sys.path.insert(0, current_dir)
sys.path.insert(0, parent_src_dir)
sys.path.insert(0, strategies_dir)

# 롤링 백테스트 설정 변수들 (여기를 수정하세요)
BATCH_CONFIG = {
    'symbol': 'S&P500',                    # 투자 지수
    'start_year': 1982,                    # 분석 시작 연도 (이 값을 변경하여 배치 조절)
    'start_month': 1,                      # 분석 시작 월 (이 값을 변경하여 배치 조절)
    'end_year': 2015,                      # 분석 종료 연도 (이 값을 변경하여 배치 조절)
    'end_month': 6,                       # 분석 종료 월 (이 값을 변경하여 배치 조절)
    'investment_period_years': 10,         # 각 테스트의 투자 기간 (년)
    'dca_months': 60,                      # 적립 분할 월수
    'generate_charts': True,               # 차트 생성 여부 (True: 생성, False: 생성 안함)
}


def run_single_backtest_silent(symbol: str, start_year: int, start_month: int, 
                              investment_period_years: int, dca_months: int) -> Dict[str, Any]:
    """단일 백테스트 실행"""
    try:
        from config import LumpSumVsDcaConfig
        from lump_sum_vs_dca_backtester import LumpSumVsDcaBacktester
        from analyzer import PerformanceAnalyzer
        
        config = LumpSumVsDcaConfig()
        config.set_analysis_params(
            symbol=symbol,
            start_year=start_year,
            start_month=start_month,
            investment_period_years=investment_period_years,
            dca_months=dca_months
        )
        
        backtester = LumpSumVsDcaBacktester(config)
        comparison_result = backtester.run_comparison(config.symbol)
        
        # 결과 추출
        lump_sum_data = comparison_result['lump_sum']['daily_returns']
        dca_data = comparison_result['dca']['daily_returns']
        
        # 기본 지표
        lump_sum_final_value = lump_sum_data['current_value'].iloc[-1]
        lump_sum_invested = lump_sum_data['invested_amount'].iloc[-1]
        lump_sum_return = (lump_sum_final_value - lump_sum_invested) / lump_sum_invested
        lump_sum_mdd = lump_sum_data['drawdown'].min()
        
        dca_final_value = dca_data['current_value'].iloc[-1]
        dca_invested = dca_data['invested_amount'].iloc[-1]
        dca_return = (dca_final_value - dca_invested) / dca_invested
        dca_mdd = dca_data['drawdown'].min()
        
        # CAGR 계산 (개별 백테스트와 동일한 방식)
        lump_sum_days = len(lump_sum_data)
        dca_days = len(dca_data)
        lump_sum_years = lump_sum_days / 365.25
        dca_years = dca_days / 365.25
        lump_sum_cagr = (lump_sum_final_value / lump_sum_invested) ** (1/lump_sum_years) - 1 if lump_sum_years > 0 else 0
        dca_cagr = (dca_final_value / dca_invested) ** (1/dca_years) - 1 if dca_years > 0 else 0
        
        # 변동성 & 샤프 (개별 백테스트와 동일한 방식)
        import numpy as np
        
        # total_return 컬럼 사용 (개별 백테스트와 동일)
        lump_sum_portfolio_returns = lump_sum_data['total_return'].diff().dropna() if 'total_return' in lump_sum_data.columns else lump_sum_data['current_value'].pct_change().dropna()
        dca_portfolio_returns = dca_data['total_return'].diff().dropna() if 'total_return' in dca_data.columns else dca_data['current_value'].pct_change().dropna()
        
        # 365.25일 기준 연환산
        lump_sum_volatility = lump_sum_portfolio_returns.std() * np.sqrt(365.25) if len(lump_sum_portfolio_returns) > 0 else 0
        dca_volatility = dca_portfolio_returns.std() * np.sqrt(365.25) if len(dca_portfolio_returns) > 0 else 0
        
        # 샤프 지수 (무위험수익률 2% 적용)
        risk_free_rate = 0.02
        lump_sum_mean_return = lump_sum_portfolio_returns.mean() * 365.25 if len(lump_sum_portfolio_returns) > 0 else 0
        dca_mean_return = dca_portfolio_returns.mean() * 365.25 if len(dca_portfolio_returns) > 0 else 0
        
        lump_sum_sharpe = (lump_sum_mean_return - risk_free_rate) / lump_sum_volatility if lump_sum_volatility > 0 else 0
        dca_sharpe = (dca_mean_return - risk_free_rate) / dca_volatility if dca_volatility > 0 else 0
        
        return {
            'start_year': start_year,
            'start_month': start_month,
            'period': f"{start_year}-{start_month:02d}",
            'end_period': f"{start_year + investment_period_years}-{start_month:02d}",
            'lump_sum_return': lump_sum_return,
            'lump_sum_cagr': lump_sum_cagr,
            'lump_sum_mdd': lump_sum_mdd,
            'lump_sum_sharpe': lump_sum_sharpe,
            'lump_sum_volatility': lump_sum_volatility,
            'lump_sum_final_value': lump_sum_final_value,
            'dca_return': dca_return,
            'dca_cagr': dca_cagr,
            'dca_mdd': dca_mdd,
            'dca_sharpe': dca_sharpe,
            'dca_volatility': dca_volatility,
            'dca_final_value': dca_final_value,
            'return_difference': lump_sum_return - dca_return,
            'cagr_difference': lump_sum_cagr - dca_cagr,
            'value_difference': lump_sum_final_value - dca_final_value,
        }
        
    except Exception:
        return None


def run_batch():
    """롤링 백테스트 실행"""
    
    # 설정 변수에서 값 가져오기
    SYMBOL = BATCH_CONFIG['symbol']
    START_YEAR = BATCH_CONFIG['start_year']
    START_MONTH = BATCH_CONFIG['start_month']
    END_YEAR = BATCH_CONFIG['end_year']
    END_MONTH = BATCH_CONFIG['end_month']
    INVESTMENT_PERIOD_YEARS = BATCH_CONFIG['investment_period_years']
    DCA_MONTHS = BATCH_CONFIG['dca_months']
    GENERATE_CHARTS = BATCH_CONFIG.get('generate_charts', True)
    
    # 전체 테스트 기간 생성
    test_periods = []
    current_date = datetime(START_YEAR, START_MONTH, 1)
    end_date = datetime(END_YEAR, END_MONTH, 1)
    
    while current_date <= end_date:
        test_periods.append((current_date.year, current_date.month))
        current_date = current_date + relativedelta(months=1)
    
    print(f"🚀 롤링 백테스트 실행")
    print(f"📊 지수: {SYMBOL}")
    print(f"📅 분석 범위: {START_YEAR}-{START_MONTH:02d} ~ {END_YEAR}-{END_MONTH:02d}")
    print(f"⏰ 투자 기간: {INVESTMENT_PERIOD_YEARS}년")
    print(f"💰 적립 분할: {DCA_MONTHS}개월")
    print(f"📋 테스트 기간: {len(test_periods)}개")
    print("-" * 60)
    
    results = []
    
    for i, (year, month) in enumerate(test_periods, 1):
        print(f"[{i:3d}] {year}-{month:02d} ~ {year + INVESTMENT_PERIOD_YEARS}-{month:02d} 테스트 중...", end=" ")
        
        result = run_single_backtest_silent(
            symbol=SYMBOL,
            start_year=year,
            start_month=month,
            investment_period_years=INVESTMENT_PERIOD_YEARS,
            dca_months=DCA_MONTHS
        )
        
        if result:
            results.append(result)
            print(f"✅ (일시:{result['lump_sum_return']:.1%}, 적립:{result['dca_return']:.1%})")
        else:
            print("❌")
    
    print("-" * 60)
    print(f"✅ 롤링 백테스트 완료: {len(results)}/{len(test_periods)}개 성공")
    
    # 결과 저장
    if results:
        # 결과 디렉토리 생성 (일시투자 vs 적립투자 결과 디렉토리)
        current_dir = Path(__file__).parent
        project_root = current_dir.parent.parent  # src -> boksl_quant 
        results_dir = project_root / "results" / "lump_sum_vs_dca"
        results_dir.mkdir(parents=True, exist_ok=True)
        
        # 파일명 생성
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"rolling_{SYMBOL}_{START_YEAR}{START_MONTH:02d}_{END_YEAR}{END_MONTH:02d}_{timestamp}.xlsx"
        filepath = results_dir / filename
        
        # DataFrame 생성 및 고급 엑셀 스타일 적용
        df = pd.DataFrame(results)
        _create_styled_excel(df, filepath, START_YEAR, END_YEAR, SYMBOL, INVESTMENT_PERIOD_YEARS, DCA_MONTHS)
        
        print(f"📊 결과 저장: {filepath}")
        
        # 요약 통계
        if len(results) > 5:
            lump_avg = df['lump_sum_return'].mean()
            dca_avg = df['dca_return'].mean()
            lump_win_rate = (df['return_difference'] > 0).mean()
            print(f"📈 요약: 일시투자 {lump_avg:.1%}, 적립투자 {dca_avg:.1%}, 일시투자 승률 {lump_win_rate:.1%}")
        
        # 인사이트 차트 생성 (설정에 따라)
        if GENERATE_CHARTS:
            print("\n📊 인사이트 차트 생성 중...")
            chart_generator = RollingChartGenerator(
                symbol=SYMBOL,
                start_year=START_YEAR,
                end_year=END_YEAR,
                investment_period_years=INVESTMENT_PERIOD_YEARS,
                dca_months=DCA_MONTHS
            )
            
            chart_files = chart_generator.generate_all_charts(results)
            print(f"📊 차트 생성 완료: {len(chart_files)}개 파일")
            for chart_name, chart_path in chart_files.items():
                print(f"  - {chart_name}: {Path(chart_path).name}")
        else:
            print("\n📊 차트 생성 건너뛰기 (설정: generate_charts=False)")
        
        return str(filepath)
    
    return None


def _create_styled_excel(df: pd.DataFrame, filepath: Path, start_year: int, end_year: int, 
                        symbol: str, investment_period_years: int, dca_months: int):
    """고급 스타일링이 적용된 엑셀 파일 생성"""
    
    # 컬럼 한글화 및 순서 정리
    df_styled = df.copy()
    
    # 불필요한 컬럼 제거 (시작년도, 시작월은 시작기간과 중복)
    columns_to_drop = ['start_year', 'start_month']
    df_styled = df_styled.drop(columns=[col for col in columns_to_drop if col in df_styled.columns])
    
    # 컬럼 순서 및 한글화
    column_mapping = {
        'period': '시작기간',
        'end_period': '종료기간',
        'lump_sum_return': '일시투자_수익률',
        'lump_sum_cagr': '일시투자_CAGR',
        'lump_sum_mdd': '일시투자_MDD',
        'lump_sum_sharpe': '일시투자_샤프지수',
        'lump_sum_volatility': '일시투자_변동성',
        'lump_sum_final_value': '일시투자_최종가치',
        'dca_return': '적립투자_수익률',
        'dca_cagr': '적립투자_CAGR',
        'dca_mdd': '적립투자_MDD',
        'dca_sharpe': '적립투자_샤프지수',
        'dca_volatility': '적립투자_변동성',
        'dca_final_value': '적립투자_최종가치',
        'return_difference': '수익률차이',
        'cagr_difference': 'CAGR차이',
        'value_difference': '가치차이'
    }
    
    df_styled = df_styled.rename(columns=column_mapping)
    
    # 워크북 생성
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 기본 시트 제거
    
    # 메인 시트 생성
    ws = wb.create_sheet("롤링백테스트결과")
    
    # 제목 추가
    title = f"{symbol} 롤링백테스트 결과 ({start_year}~{end_year}, {investment_period_years}년 투자, {dca_months}개월 적립)"
    ws['A1'] = title
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:Q1')  # 제목 셀 병합 (컬럼 수 조정: 19개 -> 17개)
    
    # 데이터 추가 (3행부터)
    for r_idx, row in enumerate(df_styled.values, 3):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    # 헤더 추가 (2행)
    for c_idx, column in enumerate(df_styled.columns, 1):
        cell = ws.cell(row=2, column=c_idx, value=column)
        cell.font = Font(bold=True)  # 개별 백테스트와 동일하게 검정색으로 수정
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 색상 정의 (개별 백테스트와 동일하게 수정)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")  # 헤더 회색
    common_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")  # 공통 연한회색
    lump_sum_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")  # 일시투자 연한파랑
    dca_fill = PatternFill(start_color="F0FFF0", end_color="F0FFF0", fill_type="solid")  # 적립투자 연한초록
    difference_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # 차이 연한주황
    
    # 테두리 스타일
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 헤더 스타일 적용
    for col in range(1, len(df_styled.columns) + 1):
        cell = ws.cell(row=2, column=col)
        cell.fill = header_fill
        cell.border = thin_border
    
    # 데이터 행 스타일 적용
    for row in range(3, len(df_styled) + 3):
        for col in range(1, len(df_styled.columns) + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            
            # 컬럼별 배경색 및 정렬 적용
            col_name = df_styled.columns[col-1]
            if col_name in ['시작기간', '종료기간']:
                cell.fill = common_fill
                # 텍스트 데이터는 중앙 정렬
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif '일시투자_' in col_name:
                cell.fill = lump_sum_fill
                # 수치 데이터는 오른쪽 정렬
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif '적립투자_' in col_name:
                cell.fill = dca_fill
                # 수치 데이터는 오른쪽 정렬
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif '차이' in col_name:
                cell.fill = difference_fill
                # 수치 데이터는 오른쪽 정렬
                cell.alignment = Alignment(horizontal='right', vertical='center')
                
            # 숫자 포맷 적용 (개별 백테스트와 동일하게 수정)
            if '수익률' in col_name or 'CAGR' in col_name or 'MDD' in col_name or '변동성' in col_name:
                cell.number_format = FORMAT_PERCENTAGE_00
                # 데이터 값 조정 제거 (개별 백테스트와 동일하게 소수값 그대로 사용)
            elif '가치' in col_name or ('차이' in col_name and col_name == '가치차이'):
                cell.number_format = "#,##0"  # 소수점 없는 금액 형식으로 수정
            elif '샤프지수' in col_name:
                cell.number_format = '0.00'  # 소수점 2자리로 수정
    
    # 열 너비 자동 조정 (불필요한 컬럼 제거)
    column_widths = {
        '시작기간': 12, '종료기간': 12,
        '일시투자_수익률': 15, '일시투자_CAGR': 15, '일시투자_MDD': 15, 
        '일시투자_샤프지수': 15, '일시투자_변동성': 15, '일시투자_최종가치': 18,
        '적립투자_수익률': 15, '적립투자_CAGR': 15, '적립투자_MDD': 15,
        '적립투자_샤프지수': 15, '적립투자_변동성': 15, '적립투자_최종가치': 18,
        '수익률차이': 12, 'CAGR차이': 12, '가치차이': 15
    }
    
    for col_idx, col_name in enumerate(df_styled.columns, 1):
        column_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = column_widths.get(col_name, 12)
    
    # 머리행 고정 (제목과 헤더 고정)
    ws.freeze_panes = 'A3'
    
    # 워크북 저장
    wb.save(filepath)


if __name__ == "__main__":
    run_batch()