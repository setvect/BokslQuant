"""
Excel 출력 모듈
"""
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1, FORMAT_PERCENTAGE_00
import os
from typing import Dict, Any
from datetime import datetime


class ExcelExporter:
    """Excel 출력 클래스"""
    
    def __init__(self, config):
        self.config = config
        self.workbook = None
        self.filename = None
    
    def export_analysis(self, comparison_result: Dict[str, Any], analyzer) -> str:
        """분석 결과를 Excel로 출력"""
        
        # 파일 경로 생성 (새로운 구조 사용)
        self.filename = self.config.get_excel_filename()
        filepath = self.config.get_excel_filepath()
        
        # 워크북 생성
        self.workbook = openpyxl.Workbook()
        
        # 기본 시트 삭제
        self.workbook.remove(self.workbook.active)
        
        # 각 시트 생성
        self._create_backtest_settings_sheet()
        self._create_purchase_history_sheet(comparison_result)
        self._create_daily_returns_sheet(comparison_result)
        self._create_analysis_summary_sheet(comparison_result, analyzer)
        
        # 파일 저장
        self.workbook.save(filepath)
        
        return filepath
    
    def _create_purchase_history_sheet(self, comparison_result: Dict[str, Any]):
        """매수 내역 시트 생성"""
        ws = self.workbook.create_sheet("매수 내역")
        
        # 헤더 설정
        headers = ['구분', '매수일', '매수가격', '매수금액', '매수수량', '누적수량', '평단가']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # 데이터 입력
        row = 2
        
        # 일시투자 데이터
        lump_sum_trades = comparison_result['lump_sum']['trades']
        cumulative_shares = 0
        cumulative_invested = 0
        
        for trade in lump_sum_trades:
            cumulative_shares += trade['shares']
            cumulative_invested += trade['amount']
            avg_price = cumulative_invested / cumulative_shares
            
            ws.cell(row=row, column=1, value="일시투자")
            ws.cell(row=row, column=2, value=trade['date'])
            ws.cell(row=row, column=3, value=trade['price'])
            ws.cell(row=row, column=4, value=trade['amount'])
            ws.cell(row=row, column=5, value=trade['shares'])
            ws.cell(row=row, column=6, value=cumulative_shares)
            ws.cell(row=row, column=7, value=avg_price)
            row += 1
        
        # 적립투자 데이터
        dca_trades = comparison_result['dca']['trades']
        cumulative_shares = 0
        cumulative_invested = 0
        
        for trade in dca_trades:
            cumulative_shares += trade['shares']
            cumulative_invested += trade['amount']
            avg_price = cumulative_invested / cumulative_shares
            
            ws.cell(row=row, column=1, value="적립투자")
            ws.cell(row=row, column=2, value=trade['date'])
            ws.cell(row=row, column=3, value=trade['price'])
            ws.cell(row=row, column=4, value=trade['amount'])
            ws.cell(row=row, column=5, value=trade['shares'])
            ws.cell(row=row, column=6, value=cumulative_shares)
            ws.cell(row=row, column=7, value=avg_price)
            row += 1
        
        # 셀 서식 적용
        self._apply_cell_formatting(ws)
    
    def _create_daily_returns_sheet(self, comparison_result: Dict[str, Any]):
        """일 수익률 변화 시트 생성"""
        ws = self.workbook.create_sheet("일 수익률 변화")
        
        # 일시투자와 적립투자 데이터 병합
        lump_sum_df = comparison_result['lump_sum']['daily_returns'].copy()
        dca_df = comparison_result['dca']['daily_returns'].copy()
        
        # 컬럼명 변경
        lump_sum_df.columns = [f'일시투자_{col}' if col != 'date' else col for col in lump_sum_df.columns]
        dca_df.columns = [f'적립투자_{col}' if col != 'date' else col for col in dca_df.columns]
        
        # 날짜 기준으로 병합
        merged_df = pd.merge(lump_sum_df, dca_df, on='date', how='outer').sort_values('date')
        
        # 컬럼 순서 정리
        columns_order = [
            'date', 
            '일시투자_price', '일시투자_invested_amount', '일시투자_shares', 
            '일시투자_average_price', '일시투자_current_value', '일시투자_total_return',
            '일시투자_peak_return', '일시투자_drawdown',
            '적립투자_invested_amount', '적립투자_shares',
            '적립투자_average_price', '적립투자_current_value', '적립투자_total_return',
            '적립투자_peak_return', '적립투자_drawdown'
        ]
        
        # 존재하는 컬럼만 선택
        available_columns = [col for col in columns_order if col in merged_df.columns]
        merged_df = merged_df[available_columns]
        
        # 컬럼명 한글화
        column_mapping = {
            'date': '날짜',
            '일시투자_price': '종가',
            '일시투자_invested_amount': '일시투자_투자금액',
            '일시투자_shares': '일시투자_보유수량',
            '일시투자_average_price': '일시투자_평균단가',
            '일시투자_current_value': '일시투자_현재가치',
            '일시투자_total_return': '일시투자_수익률',
            '일시투자_peak_return': '일시투자_전고점수익률',
            '일시투자_drawdown': '일시투자_고점대비손실폭',
            '적립투자_invested_amount': '적립투자_투자금액',
            '적립투자_shares': '적립투자_보유수량',
            '적립투자_average_price': '적립투자_평균단가',
            '적립투자_current_value': '적립투자_현재가치',
            '적립투자_total_return': '적립투자_수익률',
            '적립투자_peak_return': '적립투자_전고점수익률',
            '적립투자_drawdown': '적립투자_고점대비손실폭'
        }
        
        merged_df.rename(columns=column_mapping, inplace=True)
        
        # 데이터프레임을 시트에 입력
        for r in dataframe_to_rows(merged_df, index=False, header=True):
            ws.append(r)
        
        # 셀 서식 적용
        self._apply_cell_formatting(ws)
    
    def _create_analysis_summary_sheet(self, comparison_result: Dict[str, Any], analyzer):
        """분석 요약 시트 생성"""
        ws = self.workbook.create_sheet("분석 요약")
        
        # 지표 계산
        lump_sum_metrics = analyzer.calculate_metrics(comparison_result['lump_sum'])
        dca_metrics = analyzer.calculate_metrics(comparison_result['dca'])
        
        # 요약 테이블 생성 (순수 숫자 값으로 저장)
        summary_data = [
            ['구분', '일시투자', '적립투자', '차이'],
            ['최종 수익률', lump_sum_metrics['final_return'], 
             dca_metrics['final_return'], 
             lump_sum_metrics['final_return'] - dca_metrics['final_return']],
            ['CAGR', lump_sum_metrics['cagr'], 
             dca_metrics['cagr'], 
             lump_sum_metrics['cagr'] - dca_metrics['cagr']],
            ['MDD', lump_sum_metrics['mdd'], 
             dca_metrics['mdd'], 
             lump_sum_metrics['mdd'] - dca_metrics['mdd']],
            ['샤프 지수', lump_sum_metrics['sharpe_ratio'], 
             dca_metrics['sharpe_ratio'], 
             lump_sum_metrics['sharpe_ratio'] - dca_metrics['sharpe_ratio']],
            ['변동성', lump_sum_metrics['volatility'], 
             dca_metrics['volatility'], 
             lump_sum_metrics['volatility'] - dca_metrics['volatility']],
            ['최종 가치', lump_sum_metrics['final_value'], 
             dca_metrics['final_value'], 
             lump_sum_metrics['final_value'] - dca_metrics['final_value']]
        ]
        
        # 데이터 입력
        for row_data in summary_data:
            ws.append(row_data)
        
        # 투자 설정 정보 추가
        ws.append([])
        ws.append(['[ 투자 설정 정보 ]'])
        ws.append(['지수', self.config.symbol])
        ws.append(['투자 시작', f"{self.config.start_year}년 {self.config.start_month}월"])
        ws.append(['투자 기간', f"{self.config.investment_period_years}년"])
        ws.append(['적립 분할 월수', f"{self.config.dca_months}개월"])
        ws.append(['총 투자금', f"{self.config.initial_capital:,}"])
        ws.append(['월 적립금', f"{self.config.get_dca_monthly_amount():,.0f}"])
        
        # 셀 서식 적용
        self._apply_cell_formatting(ws)
    
    def _apply_cell_formatting(self, ws):
        """셀 서식 적용"""
        # 머리행 고정
        ws.freeze_panes = 'A2'
        
        # 색상 정의
        common_color = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")  # 연한 회색
        lump_sum_color = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")  # 연한 파란색
        dca_color = PatternFill(start_color="F0FFF0", end_color="F0FFF0", fill_type="solid")  # 연한 초록색
        
        # 테두리 정의
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 열 너비 조정 및 컬럼별 색상 적용
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            header_value = str(column[0].value) if column[0].value else ""
            
            # 컬럼 유형별 색상 결정
            if any(keyword in header_value for keyword in ['날짜', '종가', '구분', '매수일', '매수가격']):
                column_color = common_color
            elif '일시투자' in header_value:
                column_color = lump_sum_color
            elif '적립투자' in header_value:
                column_color = dca_color
            else:
                column_color = common_color
            
            # 해당 컬럼의 모든 셀에 색상 및 테두리 적용
            for cell in column:
                cell.fill = column_color
                cell.border = thin_border
                
                # 첫 번째 행(헤더)은 볼드체 및 가운데 정렬 적용
                if cell.row == 1:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center")
                
                # 길이 계산 (실제 표시될 텍스트 길이 고려)
                try:
                    cell_text = str(cell.value) if cell.value is not None else ""
                    # 한글은 2배 가중치 적용
                    text_length = sum(2 if ord(c) > 127 else 1 for c in cell_text)
                    if text_length > max_length:
                        max_length = text_length
                except:
                    pass
            
            # 데이터에 맞는 최적 너비 설정 (최소 8, 최대 25)
            optimal_width = max(8, min(max_length + 2, 25))
            ws.column_dimensions[column_letter].width = optimal_width
        
        # 데이터 타입별 서식 적용
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                
                # 숫자 타입 확인 함수 (numpy 타입 포함)
                def is_numeric(value):
                    try:
                        float(value)
                        return True
                    except (ValueError, TypeError):
                        return False
                
                # 최종 가치 행 우선 처리 (금액 형식)  
                if cell.row > 1 and ws.cell(cell.row, 1).value == '최종 가치':
                    if is_numeric(cell.value) and cell.column > 1:
                        cell.number_format = "#,##0"
                        cell.alignment = Alignment(horizontal="right")
                        continue  # 다른 조건들 건너뛰기
                
                # 날짜 형식
                elif '날짜' in str(ws.cell(1, cell.column).value) or 'date' in str(ws.cell(1, cell.column).value):
                    cell.alignment = Alignment(horizontal="center")
                
                # 수익률 형식
                elif '수익률' in str(ws.cell(1, cell.column).value) or 'return' in str(ws.cell(1, cell.column).value):
                    if is_numeric(cell.value) and cell.row > 1:
                        cell.number_format = FORMAT_PERCENTAGE_00
                        cell.alignment = Alignment(horizontal="right")
                
                # 샤프 지수 형식 (소수점 2자리)
                elif '샤프 지수' in str(ws.cell(cell.row, 1).value):
                    if is_numeric(cell.value) and cell.column > 1:
                        cell.number_format = "0.00"
                        cell.alignment = Alignment(horizontal="right")
                
                # 퍼센트 지표 형식 (CAGR, MDD, 변동성, 최종 수익률)
                elif any(keyword in str(ws.cell(cell.row, 1).value) for keyword in ['CAGR', 'MDD', '변동성', '최종 수익률']):
                    if is_numeric(cell.value) and cell.column > 1:
                        cell.number_format = FORMAT_PERCENTAGE_00
                        cell.alignment = Alignment(horizontal="right")
                
                # 금액 형식 (소수점 없음)
                elif any(keyword in str(ws.cell(1, cell.column).value) for keyword in ['금액', '가치', '가격', 'amount', 'value', 'price']):
                    if is_numeric(cell.value) and cell.row > 1:
                        cell.number_format = "#,##0"
                        cell.alignment = Alignment(horizontal="right")
                
                # 수량 및 평단가 형식 (소수점 2자리)
                elif any(keyword in str(ws.cell(1, cell.column).value) for keyword in ['수량', '평균단가', 'shares', 'average_price']):
                    if is_numeric(cell.value) and cell.row > 1:
                        cell.number_format = "#,##0.00"
                        cell.alignment = Alignment(horizontal="right")
                
                # 손실폭 형식 (소수점 2자리 퍼센트)
                elif '고점대비손실폭' in str(ws.cell(1, cell.column).value):
                    if is_numeric(cell.value) and cell.row > 1:
                        cell.number_format = "0.00%"
                        cell.alignment = Alignment(horizontal="right")
                
                # 일반 숫자 형식
                elif is_numeric(cell.value) and cell.row > 1:
                    cell.alignment = Alignment(horizontal="right")
                
                # 모든 숫자 데이터 오른쪽 정렬 (헤더 제외) - 이미 위에서 처리되지 않은 경우
                elif cell.row > 1 and is_numeric(cell.value):
                    cell.alignment = Alignment(horizontal="right")
    
    def _create_backtest_settings_sheet(self):
        """백테스트 설정 시트 생성"""
        ws = self.workbook.create_sheet("백테스트 설정")
        
        # 제목 설정
        ws.cell(row=1, column=1, value="백테스트 설정 정보")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="left")
        
        # 설정 정보 추가
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        settings_data = [
            ["실행 시간", current_time],
            ["", ""],  # 빈 행
            ["투자 설정", ""],
            ["지수", self.config.symbol],
            ["투자 시작일", f"{self.config.start_year}년 {self.config.start_month}월"],
            ["투자 기간", f"{self.config.investment_period_years}년"],
            ["적립 분할 월수", f"{self.config.dca_months}개월"],
            ["총 투자금", f"{self.config.initial_capital:,}원"],
            ["월 적립금", f"{self.config.get_dca_monthly_amount():,.0f}원"],
            ["", ""],  # 빈 행
            ["파일 정보", ""],
            ["데이터 파일", f"{self.config.symbol}_data.csv"],
            ["결과 파일", self.config.get_excel_filename()],
            ["생성 디렉토리", self.config.excel_dir],
        ]
        
        # 데이터 입력
        for row_idx, (key, value) in enumerate(settings_data, start=3):
            if key:  # 키가 있는 경우
                ws.cell(row=row_idx, column=1, value=key)
                ws.cell(row=row_idx, column=1).font = Font(bold=True)
                ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="left")
                
                if value:  # 값이 있는 경우
                    ws.cell(row=row_idx, column=2, value=value)
                    ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="left")
                    
                    # 섹션 제목 스타일링
                    if not value and key in ["투자 설정", "파일 정보"]:
                        ws.cell(row=row_idx, column=1).fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        
        # 열 너비 조정
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        
        # 머리행 고정
        ws.freeze_panes = 'A2'